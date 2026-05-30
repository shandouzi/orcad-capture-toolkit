import csv
import os
import re
import sys
import shutil
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# design by shandouzi

try:
    from tkinter import Tk, filedialog
    HAS_TK = True
except ImportError:
    HAS_TK = False


def pin_sort_key(pin):
    if not pin:
        return ("", 0)
    letter = pin[0]
    try:
        num = int(pin[1:])
    except ValueError:
        num = 0
    return (letter, num)


def parse_pin_net_filename(filepath):
    base = os.path.splitext(os.path.basename(filepath))[0]
    m = re.match(r"^(.+)_(\w+)_pin_nets$", base)
    if m:
        return m.group(1), m.group(2)
    return base, ""


def select_files_gui():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    print("Step 1/4: Select connector pin-net CSV (Board-A)")
    csv_a = filedialog.askopenfilename(
        title="Select Connector Pin-Net CSV (Board-A)",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
    )
    if not csv_a:
        print("  Cancelled.")
        sys.exit(0)
    des_a, ref_a = parse_pin_net_filename(csv_a)
    print(f"  Design: {des_a}  RefDes: {ref_a}")

    print("Step 2/4: Select connector pin-net CSV (Board-B)")
    csv_b = filedialog.askopenfilename(
        title="Select Connector Pin-Net CSV (Board-B)",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
    )
    if not csv_b:
        print("  Cancelled.")
        sys.exit(0)
    des_b, ref_b = parse_pin_net_filename(csv_b)
    print(f"  Design: {des_b}  RefDes: {ref_b}")

    print("Step 3/4: Select connector pin-map CSV")
    pin_map_csv = filedialog.askopenfilename(
        title="Select Connector Pin-Map CSV",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
    )
    if not pin_map_csv:
        print("  Cancelled.")
        sys.exit(0)
    print(f"  {pin_map_csv}")

    print("Step 4/4: Select output directory")
    output_dir = filedialog.askdirectory(title="Select Output Directory")
    if not output_dir:
        print("  Cancelled.")
        sys.exit(0)
    print(f"  {output_dir}")

    root.destroy()
    return csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b


def select_files_cli():
    print()
    print("Please provide the following files:")
    print()

    csv_a = input("  [1] Connector pin-net CSV path (Board-A): ").strip().strip('"')
    if not os.path.isfile(csv_a):
        print(f"  ERROR: File not found: {csv_a}")
        sys.exit(1)
    des_a, ref_a = parse_pin_net_filename(csv_a)
    print(f"  Design: {des_a}  RefDes: {ref_a}")

    csv_b = input("  [2] Connector pin-net CSV path (Board-B): ").strip().strip('"')
    if not os.path.isfile(csv_b):
        print(f"  ERROR: File not found: {csv_b}")
        sys.exit(1)
    des_b, ref_b = parse_pin_net_filename(csv_b)
    print(f"  Design: {des_b}  RefDes: {ref_b}")

    pin_map_csv = input("  [3] Connector pin-map CSV path: ").strip().strip('"')
    if not os.path.isfile(pin_map_csv):
        print(f"  ERROR: File not found: {pin_map_csv}")
        sys.exit(1)

    output_dir = input("  [4] Output directory (will create if not exist): ").strip().strip('"')
    if not output_dir:
        output_dir = os.path.dirname(csv_a)

    return csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b


def do_match(csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b):
    os.makedirs(output_dir, exist_ok=True)

    label_a = f"{des_a}_{ref_a}" if ref_a else des_a
    label_b = f"{des_b}_{ref_b}" if ref_b else des_b

    report_name = f"match_report_{label_a}_vs_{label_b}"
    report_csv = os.path.join(output_dir, f"{report_name}.csv")
    report_xlsx = os.path.join(output_dir, f"{report_name}.xlsx")

    col_pin_a = f"{label_a}_PinNumber"
    col_name_a = f"{label_a}_PinName"
    col_net_a = f"{label_a}_NetName"
    col_pin_b = f"{label_b}_PinNumber"
    col_name_b = f"{label_b}_PinName"
    col_net_b = f"{label_b}_NetName"

    print()
    print("Backing up source files (once) ...")
    for src in [csv_a, csv_b, pin_map_csv]:
        if os.path.exists(src):
            bak = src + ".bak"
            if not os.path.exists(bak):
                shutil.copy2(src, bak)
                print(f"  {os.path.basename(src)} -> {os.path.basename(bak)}")
            else:
                print(f"  {os.path.basename(src)}: backup already exists, skipped")

    print()
    print(f"Reading {os.path.basename(pin_map_csv)} ...")
    pin_to_mate = {}
    with open(pin_map_csv, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pin_to_mate[row["ConnectorPin"].strip()] = row["MatingPin"].strip()
    print(f"  Mappings loaded: {len(pin_to_mate)}")

    print(f"Reading {os.path.basename(csv_a)} ({label_a}) ...")
    data_a = {}
    with open(csv_a, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            data_a[row["PinNumber"].strip()] = row
    print(f"  Pins loaded: {len(data_a)}")

    print(f"Reading {os.path.basename(csv_b)} ({label_b}) ...")
    data_b = {}
    with open(csv_b, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            data_b[row["PinNumber"].strip()] = row
    print(f"  Pins loaded: {len(data_b)}")

    print()
    print(f"Matching {label_a} <-> {label_b} ...")
    results = []
    stats = {"OK": 0, "MISMATCH": 0, "NC": 0, "NC_A": 0, "NC_B": 0, "MISSING": 0}

    for pin_a in sorted(data_a.keys(), key=pin_sort_key):
        row_a = data_a[pin_a]
        net_a = row_a["NetName"].strip()
        name_a = row_a["PinName"].strip()

        mate_pin = pin_to_mate.get(pin_a)

        if mate_pin is None:
            status = "MISSING"
            pin_b = ""
            name_b = ""
            net_b = ""
        else:
            row_b = data_b.get(mate_pin)
            if row_b is None:
                status = "MISSING"
                pin_b = mate_pin
                name_b = ""
                net_b = ""
            else:
                pin_b = mate_pin
                name_b = row_b["PinName"].strip()
                net_b = row_b["NetName"].strip()

                if net_a == "" and net_b == "":
                    status = "NC"
                elif net_a == "":
                    status = "NC_A"
                elif net_b == "":
                    status = "NC_B"
                elif net_a == net_b:
                    status = "OK"
                else:
                    status = "MISMATCH"

        stats[status] += 1
        results.append({
            col_pin_a: pin_a,
            col_name_a: name_a,
            col_net_a: net_a,
            col_pin_b: pin_b,
            col_name_b: name_b,
            col_net_b: net_b,
            "Status": status,
        })

    fieldnames = [col_pin_a, col_name_a, col_net_a, col_pin_b, col_name_b, col_net_b, "Status"]
    with open(report_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Match Report"
    ws.append(fieldnames)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r in results:
        row_data = [r[col] for col in fieldnames]
        ws.append(row_data)
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        if r["Status"] == "MISMATCH":
            for cell in ws[row_num]:
                cell.fill = yellow_fill

    col_widths = [20, 26, 34, 20, 26, 34, 12]
    for i, w in enumerate(col_widths, 1):
        if i <= 7:
            ws.column_dimensions[chr(64 + i)].width = w

    wb.save(report_xlsx)

    print(f"\nReport saved:")
    print(f"  CSV:  {report_csv}")
    print(f"  XLSX: {report_xlsx}")
    print(f"Total rows: {len(results)}")
    print()
    print("=" * 40)
    print("  MATCH SUMMARY")
    print("=" * 40)
    for s in ["OK", "MISMATCH", "NC", "NC_A", "NC_B", "MISSING"]:
        if stats[s] > 0:
            print(f"  {s:<12} {stats[s]:>5}")
    print("=" * 40)

    mismatches = [r for r in results if r["Status"] == "MISMATCH"]
    if mismatches:
        print(f"\nMISMATCH details ({len(mismatches)} pins):")
        print(f"  {label_a + '_Pin':<18} {'Net':<30} {label_b + '_Pin':<18} {'Net':<30}")
        print(f"  {'-'*18} {'-'*30} {'-'*18} {'-'*30}")
        for r in mismatches:
            print(f"  {r[col_pin_a]:<18} {r[col_net_a]:<30} {r[col_pin_b]:<18} {r[col_net_b]:<30}")


def main():
    print("=" * 50)
    print("  Connector Net Matching Tool")
    print("  design by shandouzi")
    print("=" * 50)

    if HAS_TK:
        print("\nOpening file dialogs ...\n")
        csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b = select_files_gui()
    else:
        csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b = select_files_cli()

    do_match(csv_a, csv_b, pin_map_csv, output_dir, des_a, ref_a, des_b, ref_b)


if __name__ == "__main__":
    main()
